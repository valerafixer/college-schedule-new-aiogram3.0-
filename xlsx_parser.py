from openpyxl import load_workbook
from datetime import datetime

def parse_schedule_xlsx(file_path):
    """
    Парсит XLSX файл с расписанием в формате с днями недели в столбцах
    """
    try:
        wb = load_workbook(file_path)
        schedule_data = []
        
        # Маппинг дней недели
        day_mapping = {
            'понедельник': 1,
            'вторник': 2,
            'среда': 3,
            'четверг': 4,
            'пятница': 5,
            'суббота': 6,
            'воскресенье': 7
        }
        
        print(f"Найдено листов: {len(wb.worksheets)}")
        
        for sheet in wb.worksheets:
            sheet_name = sheet.title
            print(f"\n{'='*60}")
            print(f"Обрабатываю лист: {sheet_name}")
            print(f"{'='*60}")
            
            # Ищем все заголовки недель на листе
            week_headers = []
            
            for row_num in range(1, sheet.max_row + 1):
                row = list(sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
                
                for cell_value in row:
                    if cell_value:
                        cell_text = str(cell_value).lower().strip()
                        
                        # Ищем заголовки типа "Неделя 1", "Верхняя", "Нижняя"
                        if any(keyword in cell_text for keyword in ['неделя', 'верхняя', 'нижняя']):
                            week_type = None
                            if 'верхняя' in cell_text or '1' in cell_text:
                                week_type = 'upper'
                            elif 'нижняя' in cell_text or '2' in cell_text:
                                week_type = 'lower'
                            
                            if week_type:
                                week_headers.append({
                                    'type': week_type,
                                    'start_row': row_num,
                                    'name': cell_value
                                })
                                print(f"✓ Найден заголовок недели: {cell_value} (строка {row_num}, тип: {week_type})")
                                break
            
            # Если не нашли заголовки недель, обрабатываем весь лист как одну неделю
            if not week_headers:
                print("⚠ Заголовки недель не найдены, обрабатываю весь лист как верхнюю неделю")
                week_headers = [{'type': 'upper', 'start_row': 1, 'name': 'Верхняя неделя'}]
            
            # Обрабатываем каждую неделю отдельно
            for idx, week_header in enumerate(week_headers):
                week_type = week_header['type']
                start_row = week_header['start_row']
                
                # Определяем конец этой недели (начало следующей или конец листа)
                if idx + 1 < len(week_headers):
                    end_row = week_headers[idx + 1]['start_row'] - 1
                else:
                    end_row = sheet.max_row
                
                print(f"\n▶▶▶ Обрабатываю {week_header['name']} (строки {start_row}-{end_row})")
                
                # Ищем заголовок с днями недели после заголовка недели
                day_columns = {}
                header_row_num = None
                time_column = None
                para_column = None
                
                for row_num in range(start_row, min(start_row + 10, end_row + 1)):
                    row = list(sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
                    
                    for col_num, cell_value in enumerate(row, start=1):
                        if cell_value:
                            cell_text = str(cell_value).lower().strip()
                            
                            # Ищем столбец с номером пары
                            if cell_text in ['пара', '№', 'пары']:
                                para_column = col_num
                            
                            # Ищем столбец со временем
                            if 'время' in cell_text or 'врем' in cell_text:
                                time_column = col_num
                            
                            # Ищем дни недели
                            for day_name, day_num in day_mapping.items():
                                if day_name in cell_text:
                                    if col_num not in day_columns:
                                        day_columns[col_num] = day_num
                                        if header_row_num is None:
                                            header_row_num = row_num
                    
                    if day_columns:
                        break
                
                if not day_columns:
                    print(f"❌ Не найдены дни недели для {week_header['name']}")
                    continue
                
                print(f"✓ Найдено дней недели: {len(day_columns)}")
                
                # Определяем столбец с номерами пар
                lesson_num_column = para_column if para_column else time_column
                
                if lesson_num_column is None:
                    for col_num in range(1, 4):
                        if col_num not in day_columns:
                            lesson_num_column = col_num
                            break
                
                print(f"✓ Читаю с строки {header_row_num + 1} до {end_row}")
                
                # Читаем данные для этой недели
                for row_num in range(header_row_num + 1, end_row + 1):
                    row = list(sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
                    
                    # Читаем номер пары
                    lesson_cell = row[lesson_num_column - 1] if lesson_num_column and len(row) >= lesson_num_column else None
                    
                    if not lesson_cell:
                        continue
                    
                    lesson_str = str(lesson_cell).strip()
                    lesson_order = None
                    
                    # Если это просто число
                    if lesson_str.isdigit():
                        lesson_order = int(lesson_str)
                        
                        # Выводим информацию только при первом появлении номера
                        if row_num == header_row_num + 1 or lesson_order != prev_lesson:
                            time_cell = row[time_column - 1] if time_column and len(row) >= time_column else None
                            time_str = str(time_cell) if time_cell else ""
                            print(f"  📌 Пара {lesson_order}: {time_str}")
                            prev_lesson = lesson_order
                    
                    if lesson_order is None:
                        continue
                    
                    # Читаем предметы для каждого дня недели
                    for col_num, day_num in day_columns.items():
                        if col_num > len(row):
                            continue
                        
                        cell_value = row[col_num - 1]
                        
                        if cell_value:
                            subject = str(cell_value).strip()
                            subject = ' '.join(subject.split())
                            
                            if len(subject) > 3:
                                schedule_data.append((week_type, day_num, lesson_order, subject))
                
                prev_lesson = 0  # Сброс для следующей недели
        
        print(f"\n{'='*60}")
        print(f"✅ Всего загружено записей: {len(schedule_data)}")
        print(f"{'='*60}\n")
        return schedule_data
    
    except Exception as e:
        print(f"❌ Ошибка парсинга расписания: {e}")
        import traceback
        traceback.print_exc()
        return []


def parse_replacements_xlsx(file_path):
    """
    Парсит XLSX файл с заменами
    Ожидаемая структура:
    Столбцы: Дата | Текст замены
    """
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        replacements_data = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            
            date_val = row[0]
            text = str(row[1]) if row[1] else None
            
            if not text:
                continue
            
            # Обработка даты
            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            elif isinstance(date_val, str):
                try:
                    parsed_date = datetime.strptime(date_val, '%Y-%m-%d')
                    date_str = parsed_date.strftime('%Y-%m-%d')
                except:
                    try:
                        parsed_date = datetime.strptime(date_val, '%d.%m.%Y')
                        date_str = parsed_date.strftime('%Y-%m-%d')
                    except:
                        continue
            else:
                continue
            
            replacements_data.append((date_str, text))
        
        return replacements_data
    
    except Exception as e:
        print(f"Ошибка парсинга замен: {e}")
        return []
